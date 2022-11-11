# Create your tasks here
from datetime import datetime

from celery import shared_task
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from university.models import Direction_of_studying, Education_group


@shared_task
def add(x, y):
    return x + y


@shared_task
def mul(x, y):
    return x * y


@shared_task
def xsum(numbers):
    return sum(numbers)

@shared_task
def get_report_excel():
    direction_queryset = Direction_of_studying.objects.all()
    groups_queryset = Education_group.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    worksheet = workbook.create_sheet(title="Учебные направления", index=0)
    worksheet2 = workbook.create_sheet(title="Учебные группы", index=1)
    wrapped_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Define the titles for worksheet and worksheet2
    worksheet_title = [
        'ID',
        'Куратор',
        'Направление',
        'Дисциплины',
    ]

    worksheet2_title = [
        'ID',
        'Номер группы',
        'Количество свободных мест',
        'Список студентов'
    ]

    # Assign the titles for each cell of the header
    def set_worksheet_title(worksheet, title):
        for col_num, column_title in enumerate(title, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.font = Font(bold=True)

    set_worksheet_title(worksheet, worksheet_title)
    set_worksheet_title(worksheet2, worksheet2_title)

    row_num = 2
    # Iterate through all directions
    for direction in direction_queryset:

        # Define the data for each cell in the row
        try:
            curator = direction.get_curator.first_name + " " + direction.get_curator.last_name
        except ObjectDoesNotExist:
            curator = "Нет куратора"

        data = [
            direction.pk,
            curator,
            direction.direction,
        ]
        data2 = direction.get_dir.all()

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

        for col_num, cell_value in enumerate(data2, 1):
            cell = worksheet.cell(row=row_num, column=len(data) + 1)
            cell.value = cell_value.subject_name
            cell.alignment = wrapped_alignment
            row_num += 1

        if not data2:
            row_num += 1

    row_num = 2
    # Iterate through all directions
    for groups in groups_queryset:

        nums_of_students_in_group = groups.student_set.all().count()
        left_students = 20 - nums_of_students_in_group
        # Define the data for each cell in the row
        data = [
            groups.pk,
            groups.group_number,
            left_students
        ]
        data2 = groups.student_set.all().values()

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(data, 1):
            cell = worksheet2.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

        for col_num, cell_value in enumerate(data2, 1):
            cell = worksheet2.cell(row=row_num, column=len(data) + 1)
            cell.value = cell_value['first_name'] + " " + cell_value['last_name']
            cell.alignment = wrapped_alignment
            row_num += 1

        if not data2:
            row_num += 1

    def column_dimensions_width(column: int, width: int, worksheet=None):
        column_letter = get_column_letter(column)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = width

    column_dimensions_width(2, 20, worksheet)
    column_dimensions_width(3, 20, worksheet)
    column_dimensions_width(4, 40, worksheet)
    column_dimensions_width(2, 20, worksheet2)
    column_dimensions_width(3, 20, worksheet2)
    column_dimensions_width(4, 20, worksheet2)

    date = datetime.now().strftime('%Y-%m-%d')
    workbook.save(f"{date}-report.xlsx")