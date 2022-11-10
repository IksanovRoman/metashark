from datetime import datetime

from django.core.exceptions import ObjectDoesNotExist
from django.forms import model_to_dict
from django.http import HttpResponse
from django.shortcuts import render

# Create your views here.
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from rest_framework.generics import *
from rest_framework.permissions import *
from rest_framework.response import Response
from rest_framework.views import APIView

from university.models import *
from university.permissions import IsAdminOrReadOnly
from university.serializers import *


def excel(request):
    direction_queryset = Direction_of_studying.objects.all()
    groups_queryset = Education_group.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    worksheet = workbook.create_sheet(title="Учебные направления", index=0)
    worksheet2 = workbook.create_sheet(title="Учебные группы", index=1)
    wrapped_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Define the titles for worksheet and worksheet2
    worksheet_title = [
        'ID',
        'Куратор',
        'Направление',
        'Дисциплины',
    ]

    worksheet2_title = [
        'ID',
        'Номер группы',
        'Количество свободных мест',
        'Список студентов'
    ]

    # Assign the titles for each cell of the header
    def set_worksheet_title(worksheet, title):
        for col_num, column_title in enumerate(title, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.font = Font(bold=True)

    set_worksheet_title(worksheet, worksheet_title)
    set_worksheet_title(worksheet2, worksheet2_title)

    row_num = 2
    # Iterate through all directions
    for direction in direction_queryset:

        # Define the data for each cell in the row
        try:
            curator = direction.get_curator.first_name + " " + direction.get_curator.last_name
        except ObjectDoesNotExist:
            curator = "Нет куратора"

        data = [
            direction.pk,
            curator,
            direction.direction,
        ]
        data2 = direction.get_dir.all()

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

        for col_num, cell_value in enumerate(data2, 1):
            cell = worksheet.cell(row=row_num, column=len(data) + 1)
            cell.value = cell_value.subject_name
            cell.alignment = wrapped_alignment
            row_num += 1

        if not data2:
            row_num += 1

    row_num = 2
    # Iterate through all directions
    for groups in groups_queryset:

        nums_of_students_in_group = groups.student_set.all().count()
        left_students = 20 - nums_of_students_in_group
        # Define the data for each cell in the row
        data = [
            groups.pk,
            groups.group_number,
            left_students
        ]
        data2 = groups.student_set.all().values()

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(data, 1):
            cell = worksheet2.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

        for col_num, cell_value in enumerate(data2, 1):
            cell = worksheet2.cell(row=row_num, column=len(data) + 1)
            cell.value = cell_value['first_name'] + " " + cell_value['last_name']
            cell.alignment = wrapped_alignment
            row_num += 1

        if not data2:
            row_num += 1

    def column_dimensions_width(column: int, width: int, worksheet=None):
        column_letter = get_column_letter(column)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = width

    column_dimensions_width(2, 20, worksheet)
    column_dimensions_width(3, 20, worksheet)
    column_dimensions_width(4, 40, worksheet)
    column_dimensions_width(2, 20, worksheet2)
    column_dimensions_width(3, 20, worksheet2)
    column_dimensions_width(4, 20, worksheet2)

    workbook.save(response)

    return response


class DirectionAPIViewCreate(ListCreateAPIView):
    queryset = Direction_of_studying.objects.all()
    serializer_class = DirectionSerializer
    permission_classes = (IsAdminOrReadOnly,)


class DirectionAPIUpdateDestroy(RetrieveUpdateDestroyAPIView):
    queryset = Direction_of_studying.objects.all()
    serializer_class = DirectionSerializer
    permission_classes = (IsAdminUser,)


class SubjectAPIView(ListAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializerView


class SubjectAPICreate(CreateAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = (IsAdminUser,)


class SubjectAPIUpdateDestroy(RetrieveUpdateDestroyAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = (IsAdminUser,)


class StudentAPIView(ListAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializerView


class StudentAPICreate(CreateAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

    def perform_create(self, serializer):
        queryset = Student.objects.filter(studying_group=self.request.data['studying_group']).count()
        if queryset > 20:
            raise ValidationError('В этой группе 20 человек')
        serializer.save()


class StudentAPIUpdateDestroy(RetrieveUpdateDestroyAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = (IsAuthenticated,)


class GroupAPIView(ListAPIView):
    queryset = Education_group.objects.all()
    serializer_class = GroupSerializerCreateView
    permission_classes = (IsAuthenticatedOrReadOnly,)


class GroupAPICreate(CreateAPIView):
    queryset = Education_group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)


class GroupAPIUpdateDestroy(RetrieveUpdateDestroyAPIView):
    queryset = Education_group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)
